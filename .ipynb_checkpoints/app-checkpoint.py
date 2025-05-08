
import streamlit as st
import pandas as pd
import io
import os
from datetime import datetime
from openpyxl import load_workbook
from num2words import num2words
import re

# === Configuration ===
template_paths = {
    1: "PC Template-1.xlsx",
    2: "PC Template 2.xlsx",
    3: "PC Template 3.xlsx"
}
project_columns = {
    1: {1: "B"},
    2: {1: "B", 2: "E"},
    3: {1: "B", 2: "E", 3: "H"}
}
details_sheet = "DETAILS"
custom_dropdowns = {
    "Payment stage:": ["Stage Payment", "Final Payment", "Retention"],
    "Percentage of Advance payment? (as specified in the award letter)": ["0%", "25%", "40%", "50%", "60%", "70%"],
    "Is there 5% retention?": ["0%", "5%"],
    "Vat": ["0%", "7.5%"],
    "Address line 1": ["The Director", "The Chairman", "The Permanent Secretary", "The Honourable Commissioner", "The Special Adviser"]
}

# === User Mode and Context ===
st.set_page_config(page_title="Prepayment Certificate App", layout="wide")
st.title("Prepayment Certificate Filler")

user = st.sidebar.text_input("Enter Your Name (used for saving backups)", value="demo_user")
is_admin = st.sidebar.checkbox("View All Backups (Admin Only)")

# === Folder Handling ===
backup_root = "backups"
user_backup_dir = os.path.join(backup_root, user)
os.makedirs(user_backup_dir, exist_ok=True)

# === Utility Functions ===
def load_field_structure():
    df = pd.read_csv("Grouped_Field_Structure_Clean.csv")
    grouped = {}
    for _, row in df.iterrows():
        group = row['Group']
        field = (str(row['Row']), row['Label'], row.get('Options', ''))
        grouped.setdefault(group, []).append(field)
    return grouped

def load_template(project_count):
    return load_workbook(template_paths[project_count])

def write_to_details(ws, data_dict, column_map):
    for proj, entries in data_dict.items():
        col = column_map[proj]
        for row_idx, value in entries.items():
            ws[f"{col}{int(row_idx)}"] = value

def calculate_amount_due(inputs, proj, show_debug=False):
    def get(row):
        val = str(inputs.get(f"{row}_P{proj}", "0")).replace(",", "").replace("%", "").strip().lower()
        try:
            return float(val)
        except ValueError:
            return 0.0

    contract_sum = get("10")
    advance_payment_pct = get("12") / 100
    work_completed = get("13")
    retention_pct = get("14") / 100
    previous_payment = get("15")
    advance_refund_pct = get("16") / 100
    vat_pct = get("17") / 100

    advance_payment = contract_sum * advance_payment_pct
    retention = work_completed * retention_pct
    total_net_payment = work_completed - retention
    vat = total_net_payment * vat_pct
    total_net_amount = total_net_payment + vat
    advance_refund_amount = advance_refund_pct * advance_payment
    amount_due = total_net_amount - advance_refund_amount - previous_payment

    if show_debug:
        st.markdown(f"### 🧮 Debug for Project {proj}")
        st.write(f"Contract Sum: ₦{contract_sum:,.2f}")
        st.write(f"Advance: ₦{advance_payment:,.2f}")
        st.write(f"Work Completed: ₦{work_completed:,.2f}")
        st.write(f"Retention: ₦{retention:,.2f}")
        st.write(f"VAT: ₦{vat:,.2f}")
        st.write(f"Refund: ₦{advance_refund_amount:,.2f}")
        st.write(f"Previous Payment: ₦{previous_payment:,.2f}")
        st.success(f"Amount Due: ₦{amount_due:,.2f}")
    return amount_due

def amount_in_words_naira(amount):
    naira = int(amount)
    kobo = int(round((amount - naira) * 100))
    words = f"{num2words(naira, lang='en').capitalize()} naira"
    if kobo > 0:
        words += f", {num2words(kobo, lang='en')} kobo"
    return words.replace("-", " ")

def save_data_locally(all_inputs, filename=None):
    df = pd.DataFrame([all_inputs])
    df.to_csv("saved_form_data.csv", index=False)
    if filename:
        df.to_csv(os.path.join(user_backup_dir, filename), index=False)
    else:
        contractor = re.sub(r'[^\w\-]', '_', str(all_inputs.get("7_P1", "")).strip()) or "no_contractor"
        project = re.sub(r'[^\w\-]', '_', str(all_inputs.get("5_P1", "")).strip()) or "no_project"
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_name = f"{contractor}_{project}_{timestamp}.csv"
        df.to_csv(os.path.join(user_backup_dir, backup_name), index=False)

# === Start App ===
project_count = st.selectbox("Number of Projects", [1, 2, 3])
template_path = template_paths[project_count]
column_map = project_columns[project_count]
field_structure = load_field_structure()

if "restored_inputs" in st.session_state:
    all_inputs = st.session_state.pop("restored_inputs")
else:
    all_inputs = {}

# === Form Entry ===
for group, fields in field_structure.items():
    with st.expander(group, expanded=False):
        for row, label, _ in fields:
            for proj in range(1, project_count + 1):
                if proj > 1 and group in ["Date of Approval", "Address Line", "Signatories"]:
                    continue
                if proj > 1 and group == "Folio References" and label != "Inspection report File number":
                    continue

                key = f"{row}_P{proj}"
                label_suffix = f"{label} – Project {proj}" if project_count > 1 else label
                default = all_inputs.get(key, "")
                if label in custom_dropdowns:
                    all_inputs[key] = st.selectbox(label_suffix, custom_dropdowns[label], index=custom_dropdowns[label].index(default) if default in custom_dropdowns[label] else 0)
                elif row == "19":
                    continue
                elif row == "18":
                    amount = calculate_amount_due(all_inputs, proj, show_debug=True)
                    all_inputs[key] = f"{amount:,.2f}"
                    all_inputs[f"19_P{proj}"] = amount_in_words_naira(amount)
                    st.info(f"Amount Due: ₦{all_inputs[key]}")
                    st.caption(f"Amount in Words: {all_inputs[f'19_P{proj}']}")
                else:
                    all_inputs[key] = st.text_input(label_suffix, value=default)

# === Save & Download ===
contractor = all_inputs.get("7_P1", "Contractor")
project = all_inputs.get("5_P1", "Project Title")
filename = st.session_state.get("loaded_filename")

if st.button("💾 Save Offline"):
    save_data_locally(all_inputs, filename)

if st.button("📥 Download Excel"):
    wb = load_template(project_count)
    ws = wb[details_sheet]
    data_to_write = {p: {} for p in range(1, project_count + 1)}
    for key, value in all_inputs.items():
        if "_P" in key:
            row, proj = key.split("_P")
            data_to_write[int(proj)][row] = value
    write_to_details(ws, data_to_write, column_map)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    st.download_button("📂 Download Filled Excel", buffer, file_name=f"{project}_by_{contractor}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# === Backup Listing ===
st.sidebar.markdown("### 🔄 Manage Saved Backups")
all_user_dirs = [user_backup_dir] if not is_admin else [os.path.join(backup_root, u) for u in os.listdir(backup_root) if os.path.isdir(os.path.join(backup_root, u))]
all_backups = []

for user_dir in all_user_dirs:
    for f in os.listdir(user_dir):
        if f.endswith(".csv"):
            path = os.path.join(user_dir, f)
            title = f"{os.path.basename(user_dir)} | {f.replace('.csv', '').replace('_', ' ')}"
            all_backups.append((path, title))

search_term = st.sidebar.text_input("🔍 Search Backups")
filtered_backups = [b for b in all_backups if search_term.lower() in b[1].lower()]

for i, (path, title) in enumerate(filtered_backups):
    with st.sidebar.expander(title):
        col1, col2 = st.columns([2, 1])
        with col1:
            if st.button("Load", key=f"load_{i}"):
                data = pd.read_csv(path).to_dict(orient="records")[0]
                st.session_state["restored_inputs"] = data
                st.session_state["loaded_filename"] = os.path.basename(path)
                st.rerun()
        with col2:
            if st.button("🗑️", key=f"delete_{i}"):
                os.remove(path)
                st.rerun()

if st.sidebar.button("➕ Start New Blank Form"):
    if "loaded_filename" in st.session_state:
        del st.session_state["loaded_filename"]
    st.session_state["restored_inputs"] = {}
    st.rerun()
