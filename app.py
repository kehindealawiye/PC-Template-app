
import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from num2words import num2words
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
import gspread

def get_gsheet_client():
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive"
    ]
    creds_dict = st.secrets["gcp_service_account"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(creds_dict), scope)
    return gspread.authorize(creds)

def save_backup_to_gsheet(user, inputs_dict):
    try:
        gc = get_gsheet_client()
        sheet = gc.open("PC_Backups").sheet1  # Change to your actual sheet name if needed
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for k, v in inputs_dict.items():
            sheet.append_row([user, timestamp, k, v])
    except Exception as e:
        st.error(f"Failed to save to Google Sheet: {e}")


# Template paths and project column map
template_paths = {
    1: "PC Template-1.xlsx",
    2: "PC Template 2.xlsx",
    3: "PC Template 3.xlsx"
}
project_columns = {
    1: {1: "B"},
    2: {1: "B", 2: "E"},
    3: {1: "B", 2: "E", 3: "H"}
}
details_sheet = "DETAILS"

# User-defined dropdowns
custom_dropdowns = {
    "Payment stage:": ["Stage Payment", "Final Payment", "Retention"],
    "Percentage of Advance payment? (as specified in the award letter)": ["0%", "25%", "40%", "50%", "60%", "70%"],
    "Is there 5% retention?": ["0%", "5%"],
    "Vat": ["0%", "7.5%"],
    "Address line 1": ["The Director", "The Chairman", "The Permanent Secretary", "The Honourable Commissioner", "The Special Adviser"]
}

def load_field_structure():
    df = pd.read_csv("Grouped_Field_Structure_Clean.csv")
    grouped = {}
    for _, row in df.iterrows():
        group = row['Group']
        field = (str(row['Row']), row['Label'], row.get('Options', ''))
        grouped.setdefault(group, []).append(field)
    return grouped

def load_template(project_count):
    return load_workbook(template_paths[project_count])

def write_to_details(ws, data_dict, column_map):
    for proj, entries in data_dict.items():
        col = column_map[proj]
        for row_idx, value in entries.items():
            ws[f"{col}{int(row_idx)}"] = value

def get_calculated_value(advance_payment, advance_refund_pct, work_completed, retention_pct, vat_pct, previous_payment):
    base = work_completed - (retention_pct * work_completed)
    vat_amount = vat_pct * base
    advance_deduction = advance_refund_pct * advance_payment
    return base + vat_amount - advance_deduction - previous_payment

def amount_in_words_naira(amount):
    naira = int(amount)
    kobo = int(round((amount - naira) * 100))
    words = f"{num2words(naira, lang='en').capitalize()} naira"
    if kobo > 0:
        words += f", {num2words(kobo, lang='en')} kobo"
    return words.replace("-", " ")

# --- UI Styling ---
st.set_page_config(page_title="Prepayment Form", layout="wide")
st.markdown("""
    <style>
        body { background-color: #f5f7fa; }
        .stApp { padding: 2rem; }
        .st-expander {
            background-color: #ffffff !important;
            border: 1px solid #d0d0d0 !important;
            margin-bottom: 20px !important;
            border-radius: 8px !important;
        }
        .st-expander .css-1d391kg {
            background-color: #f0f4f8 !important;
        }
        h1 { color: #1a237e; }
    </style>
""", unsafe_allow_html=True)

# --- Main App Logic ---
st.title("Prepayment Certificate Filler")

project_count = st.selectbox("Number of Projects", [1, 2, 3])
template_path = template_paths[project_count]
column_map = project_columns[project_count]
field_structure = load_field_structure()
all_inputs = {}

# Form fields per group
for group, fields in field_structure.items():
    with st.expander(group, expanded=True):
        for row, label, _ in fields:
            for proj in range(1, project_count + 1):
                key = f"{row}_P{proj}"
                label_suffix = f"{label} – Project {proj}" if project_count > 1 else label
                if label == "Address line 2":
                    client_ministry = all_inputs.get(f"3_P{proj}", "")
                    all_inputs[key] = st.text_input(label_suffix, value=client_ministry, key=key)
                elif label in custom_dropdowns:
                    all_inputs[key] = st.selectbox(label_suffix, custom_dropdowns[label], key=key)
                else:
                    all_inputs[key] = st.text_input(label_suffix, key=key)

# Inspection pictures link
for proj in range(1, project_count + 1):
    key = f"link_P{proj}"
    all_inputs[key] = st.text_input(f"Link to Inspection Pictures – Project {proj}", value="https://medpicturesapp.streamlit.app/", key=key)

contractor = all_inputs.get("4_P1", "Contractor")
project_name = all_inputs.get("1_P1", "FilledTemplate")

# Excel generation
if st.button("Generate Excel"):
    wb = load_template(project_count)
    ws = wb[details_sheet]
    project_data = {p: {} for p in range(1, project_count + 1)}
    for key, value in all_inputs.items():
        if "_P" in key:
            row, proj = key.split("_P")
            project_data[int(proj)][row] = value
    write_to_details(ws, project_data, column_map)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    st.success("Excel file is ready.")
    st.download_button(
        label="Download Filled Excel",
        data=buffer,
        file_name=f"{project_name}_by_{contractor}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Calculation preview
st.subheader("Amount Due Calculation Preview")
for proj in range(1, project_count + 1):
    def parse_float(value):
        try:
            return float(str(value).replace(",", "").replace("%", ""))
        except:
            return 0.0

    work_completed = parse_float(all_inputs.get(f"11_P{proj}", 0))
    retention_pct = parse_float(all_inputs.get(f"12_P{proj}", "0")) / 100
    vat_pct = parse_float(all_inputs.get(f"13_P{proj}", "0")) / 100
    previous_payment = parse_float(all_inputs.get(f"14_P{proj}", 0))
    advance_refund_pct = parse_float(all_inputs.get(f"15_P{proj}", 0)) / 100
    advance_payment = parse_float(all_inputs.get(f"9_P{proj}", 0))

    calc_amount = get_calculated_value(
        advance_payment, advance_refund_pct,
        work_completed, retention_pct,
        vat_pct, previous_payment
    )

    words = amount_in_words_naira(calc_amount)
    st.info(f"**Project {proj} – Amount Due:** ₦{calc_amount:,.2f}")
    st.write(f"**Amount in Words:** {words}")
